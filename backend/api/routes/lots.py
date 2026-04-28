"""
Lot API routes — lot-level listing, detail, bid generation, and user actions.

Lots are the actual procurement units that suppliers bid on.
All AI analysis, profitability, and notifications operate at this level.
"""
from __future__ import annotations

import io
import math
import re
import uuid
from datetime import datetime, timezone
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, and_, or_, desc, delete as sa_delete
from sqlalchemy.ext.asyncio import AsyncSession
import httpx

from core.database import get_db
from core.deps import get_current_user
from core.subscription_guard import check_lot_view_limit
from models.user import User
from models.tender_lot import TenderLot
from models.tender import Tender
from models.tender_lot_analysis import TenderLotAnalysis
from models.profitability import ProfitabilityAnalysis
from models.logistics import LogisticsEstimate
from models.supplier import SupplierMatch, Supplier
from models.user_action import UserAction
from modules.product_resolver import resolve_product
from modules.supplier.product_search import get_product_links
from modules.supplier.product_validator import validate_products

router = APIRouter(prefix="/lots", tags=["lots"])


# ── Opportunity score ─────────────────────────────────────────────────────────

def _opportunity_score(
    margin: Optional[float],
    confidence_level: Optional[str],
    budget: Optional[float],
    deadline_at,
) -> float:
    """
    Composite score 0-100 ranking how attractive a lot is:
      margin (40%) × confidence (30%) × urgency (20%) × budget size (10%)
    Higher = better opportunity.
    """
    # Margin factor (0–1)
    if margin is None or margin < 0:
        m_factor = 0.0
    elif margin >= 40:
        m_factor = 1.0
    elif margin >= 25:
        m_factor = 0.75
    elif margin >= 15:
        m_factor = 0.5
    elif margin >= 5:
        m_factor = 0.25
    else:
        m_factor = 0.05

    # Confidence factor (0–1)
    c_map = {"high": 1.0, "medium": 0.55, "low": 0.15}
    c_factor = c_map.get(confidence_level or "", 0.0)

    # Urgency factor — deadline proximity (0–1)
    if deadline_at is None:
        u_factor = 0.3
    else:
        if hasattr(deadline_at, "tzinfo"):
            now = datetime.now(timezone.utc)
            if deadline_at.tzinfo is None:
                deadline_at = deadline_at.replace(tzinfo=timezone.utc)
            days_left = (deadline_at - now).days
        else:
            days_left = 999
        if days_left < 0:
            u_factor = 0.0
        elif days_left <= 2:
            u_factor = 1.0
        elif days_left <= 5:
            u_factor = 0.9
        elif days_left <= 10:
            u_factor = 0.7
        elif days_left <= 20:
            u_factor = 0.5
        else:
            u_factor = 0.3

    # Budget factor — log scale, capped (0–1)
    if budget and budget > 0:
        b_factor = min(1.0, math.log10(max(budget, 10_000) / 50_000) / 3.0)
        b_factor = max(0.0, b_factor)
    else:
        b_factor = 0.0

    score = m_factor * 0.40 + c_factor * 0.30 + u_factor * 0.20 + b_factor * 0.10
    return round(score * 100, 1)


def _days_until_deadline(deadline_at) -> Optional[int]:
    if deadline_at is None:
        return None
    now = datetime.now(timezone.utc)
    if deadline_at.tzinfo is None:
        deadline_at = deadline_at.replace(tzinfo=timezone.utc)
    return max(0, (deadline_at - now).days)


async def _fetch_lot_document(lot: TenderLot, tender: Tender, doc_index: int):
    documents = (lot.documents or []) + (tender.documents or [])
    if doc_index < 0 or doc_index >= len(documents):
        raise HTTPException(status_code=404, detail="Document not found")

    doc = documents[doc_index]
    doc_url = doc.get("url")
    doc_name = doc.get("name") or doc.get("filename") or f"document_{doc_index}.pdf"

    if not doc_url:
        raise HTTPException(status_code=404, detail="Document URL not available")

    if doc_url.startswith("//"):
        host = "https://goszakup.gov.kz" if lot.platform == "goszakup" else "https://zakup.sk.kz"
        doc_url = f"{host}{doc_url}"
    elif doc_url.startswith("/"):
        host = "https://goszakup.gov.kz" if lot.platform == "goszakup" else "https://zakup.sk.kz"
        doc_url = f"{host}{doc_url}"

    try:
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            response = await client.get(doc_url)
            response.raise_for_status()
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Failed to download document: {str(exc)}")

    content_type = response.headers.get("content-type") or "application/octet-stream"
    encoded_filename = quote(doc_name, safe='')
    return response.content, content_type, encoded_filename


# ── List lots ─────────────────────────────────────────────────────────────────

@router.get("", response_model=dict)
async def list_lots(
    platform: Optional[str] = None,
    category: Optional[str] = None,
    is_profitable: Optional[bool] = None,
    confidence_level: Optional[str] = None,
    min_budget: Optional[float] = None,
    max_budget: Optional[float] = None,
    search: Optional[str] = None,
    notification_sent: Optional[bool] = None,
    min_accuracy: Optional[int] = Query(None, ge=0, le=100),
    only_analyzed: Optional[bool] = Query(None),
    new_today: Optional[bool] = Query(None, description="Only show lots added today"),
    sort_by: Optional[str] = Query(None, pattern="^(newest|deadline|budget|margin|profit)$"),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    List tender lots with filters and pagination.
    Returns lots ordered by first_seen_at descending (newest first).
    Includes expected_profit and tender external_id for full row display.
    """
    from sqlalchemy import cast, Date, func as sqlfunc

    # Build ordering
    from datetime import date
    if sort_by == "deadline":
        order_clause = TenderLot.deadline_at.asc().nullslast()
    elif sort_by == "budget":
        order_clause = desc(TenderLot.budget)
    elif sort_by == "margin":
        order_clause = desc(ProfitabilityAnalysis.profit_margin_percent).nullslast()
    elif sort_by == "profit":
        order_clause = desc(ProfitabilityAnalysis.expected_profit).nullslast()
    else:
        order_clause = desc(TenderLot.first_seen_at)

    query = (
        select(TenderLot, Tender, ProfitabilityAnalysis, TenderLotAnalysis)
        .join(Tender, TenderLot.tender_id == Tender.id)
        .outerjoin(
            ProfitabilityAnalysis,
            ProfitabilityAnalysis.lot_id == TenderLot.id,
        )
        .outerjoin(
            TenderLotAnalysis,
            TenderLotAnalysis.lot_id == TenderLot.id,
        )
        .order_by(order_clause)
    )

    conditions = []
    if platform:
        conditions.append(TenderLot.platform == platform)
    if category:
        conditions.append(TenderLot.category == category)
    if is_profitable is not None:
        conditions.append(TenderLot.is_profitable == is_profitable)
    if confidence_level:
        conditions.append(TenderLot.confidence_level == confidence_level)
    if min_budget:
        conditions.append(TenderLot.budget >= min_budget)
    if max_budget:
        conditions.append(TenderLot.budget <= max_budget)
    if notification_sent is not None:
        conditions.append(TenderLot.notification_sent == notification_sent)
    if min_accuracy is not None:
        # Filter by confidence_score ≥ min_accuracy/100 (requires profitability analysis to exist)
        conditions.append(ProfitabilityAnalysis.confidence_score >= min_accuracy / 100.0)
    if only_analyzed:
        # Only show lots that have been analyzed (have profitability data)
        conditions.append(ProfitabilityAnalysis.id.isnot(None))
    if new_today:
        from datetime import date, timezone as tz_module
        today_start = datetime.combine(date.today(), datetime.min.time()).replace(tzinfo=timezone.utc)
        conditions.append(TenderLot.first_seen_at >= today_start)
    if search:
        conditions.append(
            or_(
                TenderLot.title.ilike(f"%{search}%"),
                TenderLot.description.ilike(f"%{search}%"),
                Tender.customer_name.ilike(f"%{search}%"),
            )
        )

    if conditions:
        query = query.where(and_(*conditions))

    # Total count
    count_query = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_query)).scalar()

    # Paginated results
    offset = (page - 1) * per_page
    rows = (await db.execute(query.offset(offset).limit(per_page))).all()

    items = []
    for lot, tender, prof, analysis in rows:
        _margin = float(prof.profit_margin_percent) if prof and prof.profit_margin_percent is not None else None
        # Profit label thresholds calibrated for realistic KZ procurement margins (15-40%)
        if _margin is None:
            profit_label = "unknown"
        elif _margin < 0:
            profit_label = "loss"
        elif _margin >= 30:   # 30%+ → high (was 40%)
            profit_label = "high"
        elif _margin >= 15:   # 15-30% → medium (was 20%)
            profit_label = "medium"
        else:
            profit_label = "low"

        # Accuracy: confidence_score → percent; suspicious flag
        _conf_score = float(prof.confidence_score) if prof and prof.confidence_score is not None else None
        accuracy_pct = round(_conf_score * 100, 1) if _conf_score is not None else None

        # "Suspicious" if margin > 40% but confidence is low
        _conf_level = lot.confidence_level
        is_suspicious = (
            _margin is not None
            and _margin > 40.0
            and _conf_level == "low"
        )

        items.append({
            "id": str(lot.id),
            "tender_id": str(lot.tender_id),
            "platform": lot.platform,
            "tender_external_id": tender.external_id,
            "lot_external_id": lot.lot_external_id,
            "title": lot.title,
            "category": lot.category,
            "budget": float(lot.budget) if lot.budget else None,
            "currency": lot.currency,
            "quantity": float(lot.quantity) if lot.quantity else None,
            "unit": lot.unit,
            "status": lot.status,
            "deadline_at": lot.deadline_at.isoformat() if lot.deadline_at else None,
            "first_seen_at": lot.first_seen_at.isoformat() if lot.first_seen_at else None,
            "is_profitable": lot.is_profitable,
            "profit_margin_percent": _margin,
            "confidence_level": _conf_level,
            "expected_profit": float(prof.expected_profit) if prof and prof.expected_profit else None,
            "profit_label": profit_label,
            "accuracy_pct": accuracy_pct,
            "is_suspicious": is_suspicious,
            "is_analyzed": lot.is_analyzed,
            "notification_sent": lot.notification_sent,
            "product_name": analysis.product_name if analysis else None,
            "characteristics": analysis.characteristics if analysis else None,
            "ai_summary_ru": analysis.ai_summary_ru if analysis else None,
            "spec_clarity": analysis.spec_clarity if analysis else None,
            "customer_name": tender.customer_name,
            "customer_region": tender.customer_region,
            "tender_title": tender.title,
            "opportunity_score": _opportunity_score(_margin, _conf_level, float(lot.budget) if lot.budget else None, lot.deadline_at),
            "days_until_deadline": _days_until_deadline(lot.deadline_at),
        })

    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page if total else 0,
    }


# ── Top opportunities (for dashboard widget) ──────────────────────────────────

@router.get("/top")
async def get_top_opportunities(
    limit: int = Query(8, ge=1, le=20),
    min_margin: float = Query(10.0, ge=0),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Return top N most attractive lots ranked by opportunity score.
    Used for the dashboard 'Hot Lots' widget.
    Only includes profitable, analyzed, non-expired lots.
    """
    now = datetime.now(timezone.utc)
    query = (
        select(TenderLot, Tender, ProfitabilityAnalysis)
        .join(Tender, TenderLot.tender_id == Tender.id)
        .join(ProfitabilityAnalysis, ProfitabilityAnalysis.lot_id == TenderLot.id)
        .where(
            TenderLot.is_profitable == True,  # noqa
            TenderLot.is_analyzed == True,    # noqa
            ProfitabilityAnalysis.profit_margin_percent >= min_margin,
            or_(
                TenderLot.deadline_at.is_(None),
                TenderLot.deadline_at > now,
            ),
        )
        .order_by(desc(TenderLot.first_seen_at))
        .limit(200)   # Score in Python (no SQL formula needed)
    )
    rows = (await db.execute(query)).all()

    scored = []
    for lot, tender, prof in rows:
        _margin = float(prof.profit_margin_percent) if prof.profit_margin_percent else None
        _conf   = lot.confidence_level
        _budget = float(lot.budget) if lot.budget else None
        score   = _opportunity_score(_margin, _conf, _budget, lot.deadline_at)
        days    = _days_until_deadline(lot.deadline_at)

        scored.append({
            "id":                    str(lot.id),
            "tender_id":             str(lot.tender_id),
            "platform":              lot.platform,
            "tender_external_id":    tender.external_id,
            "title":                 lot.title,
            "category":              lot.category,
            "budget":                _budget,
            "profit_margin_percent": _margin,
            "expected_profit":       float(prof.expected_profit) if prof.expected_profit else None,
            "confidence_level":      _conf,
            "opportunity_score":     score,
            "days_until_deadline":   days,
            "deadline_at":           lot.deadline_at.isoformat() if lot.deadline_at else None,
            "first_seen_at":         lot.first_seen_at.isoformat() if lot.first_seen_at else None,
            "customer_name":         tender.customer_name,
            "customer_region":       tender.customer_region,
        })

    # Sort by score DESC and return top N
    scored.sort(key=lambda x: x["opportunity_score"], reverse=True)
    return {"items": scored[:limit], "total": len(scored)}


# ── Excel export ─────────────────────────────────────────────────────────────

@router.get("/export")
async def export_lots_excel(
    platform: Optional[str] = None,
    category: Optional[str] = None,
    is_profitable: Optional[bool] = None,
    confidence_level: Optional[str] = None,
    min_budget: Optional[float] = None,
    max_budget: Optional[float] = None,
    search: Optional[str] = None,
    min_accuracy: Optional[int] = Query(None, ge=0, le=100),
    only_analyzed: Optional[bool] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Export filtered lots to Excel (.xlsx).
    Applies the same filters as the lots list. Returns up to 5 000 rows.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    query = (
        select(TenderLot, Tender, ProfitabilityAnalysis, TenderLotAnalysis)
        .join(Tender, TenderLot.tender_id == Tender.id)
        .outerjoin(ProfitabilityAnalysis, ProfitabilityAnalysis.lot_id == TenderLot.id)
        .outerjoin(TenderLotAnalysis, TenderLotAnalysis.lot_id == TenderLot.id)
        .order_by(desc(TenderLot.first_seen_at))
        .limit(5000)
    )

    conditions = []
    if platform:
        conditions.append(TenderLot.platform == platform)
    if category:
        conditions.append(TenderLot.category == category)
    if is_profitable is not None:
        conditions.append(TenderLot.is_profitable == is_profitable)
    if confidence_level:
        conditions.append(TenderLot.confidence_level == confidence_level)
    if min_budget:
        conditions.append(TenderLot.budget >= min_budget)
    if max_budget:
        conditions.append(TenderLot.budget <= max_budget)
    if min_accuracy is not None:
        conditions.append(ProfitabilityAnalysis.confidence_score >= min_accuracy / 100.0)
    if only_analyzed:
        conditions.append(ProfitabilityAnalysis.id.isnot(None))
    if search:
        conditions.append(
            or_(
                TenderLot.title.ilike(f"%{search}%"),
                Tender.customer_name.ilike(f"%{search}%"),
            )
        )
    if conditions:
        query = query.where(and_(*conditions))

    rows = (await db.execute(query)).all()

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Тендеры"

    # Header style
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1E3A5F")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        ("№",                  6),
        ("Платформа",         12),
        ("Номер тендера",     16),
        ("Наименование лота", 40),
        ("Товар (AI)",        30),
        ("Заказчик",          25),
        ("Регион",            15),
        ("Бюджет (₸)",        16),
        ("Маржа %",           10),
        ("Прибыль (₸)",       16),
        ("Точность %",        12),
        ("Уверенность",       14),
        ("Прибыльный",        12),
        ("Дедлайн",           13),
        ("Дата находки",      13),
        ("Статус",            12),
        ("Категория",         14),
    ]

    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Fill styles
    fill_profit = PatternFill("solid", fgColor="E8F5E9")  # light green
    fill_loss   = PatternFill("solid", fgColor="FFEBEE")  # light red
    fill_medium = PatternFill("solid", fgColor="FFF9C4")  # light yellow

    def fmt_num(v):
        return round(v, 2) if v is not None else ""

    def fmt_date(d):
        return d.strftime("%d.%m.%Y") if d else ""

    for row_idx, (lot, tender, prof, analysis) in enumerate(rows, start=2):
        margin = float(prof.profit_margin_percent) if prof and prof.profit_margin_percent is not None else None
        profit = float(prof.expected_profit) if prof and prof.expected_profit is not None else None
        acc    = round(float(prof.confidence_score) * 100, 1) if prof and prof.confidence_score is not None else None

        # Row fill based on profitability
        if lot.is_profitable:
            row_fill = fill_profit
        elif margin is not None and margin < 0:
            row_fill = fill_loss
        else:
            row_fill = fill_medium

        values = [
            row_idx - 1,
            lot.platform or "",
            tender.external_id or "",
            lot.title or "",
            (analysis.product_name if analysis else "") or "",
            tender.customer_name or "",
            tender.customer_region or "",
            fmt_num(float(lot.budget) if lot.budget else None),
            fmt_num(margin),
            fmt_num(profit),
            acc if acc is not None else "",
            lot.confidence_level or "",
            "Да" if lot.is_profitable else ("Нет" if lot.is_profitable is not None else ""),
            fmt_date(lot.deadline_at),
            fmt_date(lot.first_seen_at),
            lot.status or "",
            lot.category or "",
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill   = row_fill
            if col_idx in (8, 10):  # budget, profit — number format
                cell.number_format = '#,##0.00'
            if col_idx == 9:  # margin %
                cell.number_format = '0.00"%"'
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")

    # Add autofilter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # ── Stream response ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename  = f"tenders_export_{timestamp}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Lot detail ────────────────────────────────────────────────────────────────

@router.get("/{lot_id}")
async def get_lot(
    lot_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(check_lot_view_limit),
):
    """
    Full lot detail: AI analysis, profitability breakdown, suppliers, logistics.
    """
    row = await db.execute(
        select(TenderLot, Tender)
        .join(Tender, TenderLot.tender_id == Tender.id)
        .where(TenderLot.id == lot_id)
    )
    result = row.one_or_none()
    if not result:
        raise HTTPException(status_code=404, detail="Lot not found")

    lot, tender = result

    # Auto-extract spec text on first open if not yet populated (no full analysis needed)
    if not (lot.technical_spec_text or "").strip():
        try:
            import asyncio as _aio
            tech_text, raw_text, pdf_url = await _aio.wait_for(
                _refresh_spec_text(lot, tender, force=False),
                timeout=30.0,
            )
            updated = False
            if tech_text and len(tech_text.strip()) > 50:
                lot.technical_spec_text = tech_text
                lot.raw_spec_text = raw_text
                updated = True
            if pdf_url and not lot.techspec_pdf_url:
                lot.techspec_pdf_url = pdf_url
                updated = True
            if updated:
                await db.commit()
        except Exception as _exc:
            print(f"[auto spec] failed for lot {lot_id}: {_exc}", flush=True)

    # AI analysis
    analysis_row = await db.execute(
        select(TenderLotAnalysis).where(TenderLotAnalysis.lot_id == lot_id)
    )
    analysis = analysis_row.scalar_one_or_none()

    # Profitability
    prof_row = await db.execute(
        select(ProfitabilityAnalysis)
        .where(ProfitabilityAnalysis.lot_id == lot_id)
        .order_by(desc(ProfitabilityAnalysis.created_at))
        .limit(1)
    )
    prof = prof_row.scalar_one_or_none()

    # Logistics
    log_row = await db.execute(
        select(LogisticsEstimate)
        .where(LogisticsEstimate.lot_id == lot_id)
        .order_by(desc(LogisticsEstimate.created_at))
        .limit(1)
    )
    logistics = log_row.scalar_one_or_none()

    # Supplier matches
    sup_rows = await db.execute(
        select(SupplierMatch, Supplier)
        .outerjoin(Supplier, SupplierMatch.supplier_id == Supplier.id)
        .where(SupplierMatch.lot_id == lot_id)
        .order_by(SupplierMatch.match_score.desc())
    )
    suppliers = [
        {
            "supplier_name": s.name if s else None,
            "country": s.country if s else None,
            "source": s.source if s else None,
            "unit_price_kzt": float(sm.unit_price_kzt) if sm.unit_price_kzt else None,
            "unit_price_usd": float(sm.unit_price) if sm.unit_price else None,
            "lead_time_days": sm.lead_time_days,
            "match_score": float(sm.match_score) if sm.match_score else None,
            "source_url": sm.source_url,
            "contact_info": s.contact_info if s else {},
        }
        for sm, s in sup_rows.all()
    ]

    # Structured product resolution (spec text → name + model + standard)
    resolved = resolve_product(
        spec_text=lot.technical_spec_text or "",
        title=lot.title or "",
        ai_product_name=(analysis.product_name or "") if analysis else "",
        ai_brand_model=(analysis.brand_model or "") if analysis else "",
        ai_brand=(analysis.brand or "") if analysis else "",
        ai_technical_params=(analysis.technical_params or {}) if analysis else {},
    )

    # If resolver found no model, fall back to brand_model from AI analysis
    if not resolved.get("model") and analysis and analysis.brand_model:
        resolved["model"] = analysis.brand_model

    # Attach AI-extracted characteristics to resolved_product
    _chars_from_db = (analysis.characteristics or None) if analysis else None
    print(
        f"[API /lots/{lot_id}] characteristics from DB: {_chars_from_db!r}",
        flush=True,
    )
    resolved["characteristics"] = _chars_from_db

    # Attach AI suggestion fields so the frontend has one source of truth (resolved_product).
    # Normalized columns (suggested_model, suggestion_confidence) are preferred;
    # fall back to raw_ai_response for lots analyzed before those columns were added.
    _raw = (analysis.raw_ai_response or {}) if analysis else {}
    suggested_model = (
        (analysis.suggested_model if analysis else None)
        or _raw.get("suggested_model")
    ) or None  # collapse empty strings to None

    _conf_raw = (
        (analysis.suggestion_confidence if analysis else None)
        if (analysis and analysis.suggestion_confidence is not None)
        else _raw.get("suggestion_confidence")
    )

    if suggested_model and _conf_raw is not None:
        try:
            c = float(_conf_raw)
            if c <= 1.0:
                c *= 100
            confidence: int | None = max(0, min(100, int(round(c))))
        except (TypeError, ValueError):
            confidence = None
    elif suggested_model:
        confidence = None
    else:
        suggested_model = None
        confidence = None

    resolved["suggested_model"] = suggested_model
    resolved["confidence"]      = confidence

    print(
        f"[API /lots] FINAL characteristics going to frontend: {_chars_from_db!r}",
        flush=True,
    )

    # ── Marketplace links (real product search URLs across KZ/RU/CN) ─────────
    # Use enriched search_query from product resolver (includes materials, description).
    _product_name_for_links = resolved["search_query"] or (
        (analysis.product_name if analysis and analysis.product_name else None)
        or lot.title
        or ""
    )
    _tech_params_for_links = (analysis.technical_params or {}) if analysis else {}
    _product_name_en = (analysis.product_name_en or "") if analysis else ""

    marketplace_links: list[dict] = []
    try:
        import asyncio as _asyncio
        marketplace_links = await _asyncio.wait_for(
            get_product_links(
                product_name=_product_name_for_links,
                characteristics=_tech_params_for_links,
                product_name_en=_product_name_en,
                max_links=8,
            ),
            timeout=9.0,
        )
    except Exception:
        pass

    # GPT-validate real product pages against the spec (cache-first, fast)
    if marketplace_links:
        try:
            import asyncio as _asyncio
            _spec_text = (lot.technical_spec_text or lot.description or "")[:600]
            marketplace_links = await _asyncio.wait_for(
                validate_products(
                    product_name=_product_name_for_links,
                    characteristics=_tech_params_for_links,
                    products=marketplace_links,
                    spec_text=_spec_text,
                ),
                timeout=18.0,
            )
        except Exception:
            pass

    # Attach marketplace_links to each supplier row so the frontend has it
    for sup in suppliers:
        sup["marketplace_links"] = marketplace_links

    # ── Buy recommendation (based on profitability) ───────────────────────────
    _margin = float(prof.profit_margin_percent) if prof and prof.profit_margin_percent is not None else None
    _expected_profit = float(prof.expected_profit) if prof and prof.expected_profit is not None else None
    if _margin is None:
        buy_recommendation = {
            "label":  "⏳ Анализ не выполнен",
            "detail": "Запустите «Пересчитать маржу» для получения рекомендации.",
            "level":  "unknown",
        }
    elif _margin < 0:
        _loss_kzt = abs(_expected_profit) if _expected_profit is not None else 0
        buy_recommendation = {
            "label":  "❌ УБЫТОЧНЫЙ ЛОТ — участие принесёт убыток",
            "detail": f"Расчётный убыток: {_loss_kzt:,.0f} ₸. Себестоимость товара превышает бюджет лота.",
            "level":  "loss",
        }
    elif _margin >= 40:
        buy_recommendation = {
            "label":  "🔥 Выгодный тендер — рекомендуется участвовать",
            "detail": f"Маржа {_margin:.1f}% превышает целевой порог 40%.",
            "level":  "high",
        }
    elif _margin >= 20:
        buy_recommendation = {
            "label":  "⚠️ Средняя маржа — требует проверки цены",
            "detail": f"Маржа {_margin:.1f}%. Уточните цену поставщика перед участием.",
            "level":  "medium",
        }
    else:
        buy_recommendation = {
            "label":  "❌ Низкая маржа — не рекомендуется",
            "detail": f"Маржа {_margin:.1f}% ниже порога рентабельности 20%.",
            "level":  "low",
        }

    # ── Identification confidence (0–100, based on spec quality) ─────────────
    _id_score = 0
    if analysis:
        _clarity_map = {"clear": 50, "partial": 30, "vague": 10}
        _id_score += _clarity_map.get(analysis.spec_clarity or "vague", 10)
        if analysis.brand:                                      _id_score += 15
        if analysis.brand_model:                                _id_score += 15
        if analysis.characteristics:                            _id_score += 10
        if analysis.key_requirements:                           _id_score += 5
        if lot.technical_spec_text and len(lot.technical_spec_text) > 200: _id_score += 5
    identification_confidence = min(100, _id_score)

    return {
        "id": str(lot.id),
        "tender_id": str(lot.tender_id),
        "platform": lot.platform,
        "lot_external_id": lot.lot_external_id,
        "title": lot.title,
        "description": lot.description,
        "technical_spec_text": lot.technical_spec_text,
        "raw_spec_text": lot.raw_spec_text,
        "techspec_pdf_url": getattr(lot, "techspec_pdf_url", None),
        # Top-level shortcut — frontend reads this directly (no nesting needed)
        "characteristics": _chars_from_db,
        # Purchase recommendation + identification quality
        "buy_recommendation":        buy_recommendation,
        "identification_confidence": identification_confidence,
        # Structured product identity — includes strict model + AI suggestion
        "resolved_product": resolved,
        "category": lot.category,
        "budget": float(lot.budget) if lot.budget else None,
        "currency": lot.currency,
        "quantity": float(lot.quantity) if lot.quantity else None,
        "unit": lot.unit,
        "status": lot.status,
        "deadline_at": lot.deadline_at.isoformat() if lot.deadline_at else None,
        "first_seen_at": lot.first_seen_at.isoformat() if lot.first_seen_at else None,
        "documents": lot.documents,
        "tender_documents": tender.documents,
        "is_analyzed": lot.is_analyzed,
        "is_profitable": lot.is_profitable,
        "profit_margin_percent": float(lot.profit_margin_percent) if lot.profit_margin_percent else None,
        "confidence_level": lot.confidence_level,
        "notification_sent": lot.notification_sent,
        # Parent tender
        "tender": {
            "id": str(tender.id),
            "title": tender.title,
            "external_id": tender.external_id,
            "customer_name": tender.customer_name,
            "customer_bin": tender.customer_bin,
            "customer_region": tender.customer_region,
            "procurement_method": tender.procurement_method,
            "published_at": tender.published_at.isoformat() if tender.published_at else None,
        },
        # AI analysis
        "analysis": {
            "product_name": analysis.product_name,
            "product_name_en": analysis.product_name_en,
            "brand": analysis.brand,
            "brand_model": analysis.brand_model,
            "characteristics": analysis.characteristics,
            "product_type":        (_raw.get("product_type") or None),
            "normalized_name":     (_raw.get("normalized_name") or None),
            "exact_product_match": (_raw.get("exact_product_match") or None),
            "key_specs":           (_raw.get("key_specs") or []),
            "procurement_hint":    (_raw.get("procurement_hint") or None),
            "is_standard_based":   (_raw.get("is_standard_based") or False),
            "possible_suppliers":  (_raw.get("possible_suppliers") or []),
            "dimensions": analysis.dimensions,
            "technical_params": analysis.technical_params,
            "materials": analysis.materials,
            "quantity_extracted": float(analysis.quantity_extracted) if analysis.quantity_extracted else None,
            "unit_extracted": analysis.unit_extracted,
            "analogs_allowed": analysis.analogs_allowed,
            "spec_clarity": analysis.spec_clarity,
            "key_requirements": analysis.key_requirements,
            "ai_summary_ru": analysis.ai_summary_ru,
            "is_software_related": analysis.is_software_related,
            "software_type": analysis.software_type,
            "extraction_confidence": float(analysis.extraction_confidence) if analysis.extraction_confidence else None,
            "analyzed_at": analysis.analyzed_at.isoformat() if analysis.analyzed_at else None,
        } if analysis else None,
        # Profitability
        "profitability": {
            "product_cost": float(prof.product_cost) if prof.product_cost else None,
            "logistics_cost": float(prof.logistics_cost) if prof.logistics_cost else None,
            "customs_cost": float(prof.customs_cost) if prof.customs_cost else None,
            "vat_amount": float(prof.vat_amount) if prof.vat_amount else None,
            "operational_costs": float(prof.operational_costs) if prof.operational_costs else None,
            "total_cost": float(prof.total_cost) if prof.total_cost else None,
            "expected_profit": float(prof.expected_profit) if prof.expected_profit else None,
            "profit_margin_percent": float(prof.profit_margin_percent) if prof.profit_margin_percent else None,
            "is_profitable": prof.is_profitable,
            "confidence_level": prof.confidence_level,
            "confidence_score": float(prof.confidence_score) if prof.confidence_score else None,
            "recommended_bid": float(prof.recommended_bid) if prof.recommended_bid else None,
            "safe_bid": float(prof.safe_bid) if prof.safe_bid else None,
            "aggressive_bid": float(prof.aggressive_bid) if prof.aggressive_bid else None,
            "risk_level": prof.risk_level,
            "origin_country": prof.origin_country,
        } if prof else None,
        # Logistics
        "logistics": {
            "origin_country": logistics.origin_country,
            "shipping_cost": float(logistics.shipping_cost) if logistics.shipping_cost else None,
            "customs_duty": float(logistics.customs_duty) if logistics.customs_duty else None,
            "vat_amount": float(logistics.vat_amount) if logistics.vat_amount else None,
            "total_logistics": float(logistics.total_logistics) if logistics.total_logistics else None,
            "lead_time_days": logistics.lead_time_days,
            "route": logistics.route,
        } if logistics else None,
        "suppliers": suppliers,
        "marketplace_links": marketplace_links,  # available even when suppliers list is empty
    }


# ── Per-lot re-analysis ───────────────────────────────────────────────────────

@router.get("/{lot_id}/download/{doc_index}")
async def download_lot_document(
    lot_id: uuid.UUID,
    doc_index: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Download a document attached to this lot or to its parent tender.
    """
    row = await db.execute(
        select(TenderLot, Tender)
        .join(Tender, TenderLot.tender_id == Tender.id)
        .where(TenderLot.id == lot_id)
    )
    result = row.one_or_none()
    if not result:
        raise HTTPException(status_code=404, detail="Lot not found")

    lot, tender = result
    documents = (lot.documents or []) + (tender.documents or [])
    if doc_index < 0 or doc_index >= len(documents):
        raise HTTPException(status_code=404, detail="Document not found")

    doc = documents[doc_index]
    doc_url = doc.get("url")
    doc_name = doc.get("name") or doc.get("filename") or f"document_{doc_index}.pdf"

    if not doc_url:
        raise HTTPException(status_code=404, detail="Document URL not available")

    if doc_url.startswith("//"):
        host = "https://goszakup.gov.kz" if lot.platform == "goszakup" else "https://zakup.sk.kz"
        doc_url = f"{host}{doc_url}"
    elif doc_url.startswith("/"):
        host = "https://goszakup.gov.kz" if lot.platform == "goszakup" else "https://zakup.sk.kz"
        doc_url = f"{host}{doc_url}"

    try:
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            response = await client.get(doc_url)
            response.raise_for_status()
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Failed to download document: {str(exc)}")

    content_type = response.headers.get("content-type") or "application/octet-stream"
    encoded_filename = quote(doc_name, safe='')
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
    }
    if response.headers.get("content-length"):
        headers["Content-Length"] = response.headers.get("content-length")

    return StreamingResponse(
        io.BytesIO(response.content),
        media_type=content_type,
        headers=headers,
    )


@router.get("/{lot_id}/view/{doc_index}")
async def view_lot_document(
    lot_id: uuid.UUID,
    doc_index: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    View a document attached to this lot or its parent tender inline.
    """
    row = await db.execute(
        select(TenderLot, Tender)
        .join(Tender, TenderLot.tender_id == Tender.id)
        .where(TenderLot.id == lot_id)
    )
    result = row.one_or_none()
    if not result:
        raise HTTPException(status_code=404, detail="Lot not found")

    lot, tender = result
    content, content_type, encoded_filename = await _fetch_lot_document(lot, tender, doc_index)

    headers = {
        "Content-Disposition": f"inline; filename*=UTF-8''{encoded_filename}"
    }
    if content:
        headers["Content-Length"] = str(len(content))

    return StreamingResponse(
        io.BytesIO(content),
        media_type=content_type,
        headers=headers,
    )


@router.post("/{lot_id}/reanalyze")
async def reanalyze_lot(
    lot_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Force re-analysis of a single lot with the current AI pipeline.
    Deletes the existing TenderLotAnalysis record and re-runs ai_analysis_step.
    """
    from modules.scanner.pipeline import PipelineContext
    from modules.ai_analyzer.pipeline_step import ai_analysis_step

    row = await db.execute(
        select(TenderLot, Tender)
        .join(Tender, TenderLot.tender_id == Tender.id)
        .where(TenderLot.id == lot_id)
    )
    result = row.one_or_none()
    if not result:
        raise HTTPException(status_code=404, detail="Lot not found")
    lot, tender = result

    # Delete old analysis so pipeline_step inserts a fresh one
    await db.execute(sa_delete(TenderLotAnalysis).where(TenderLotAnalysis.lot_id == lot_id))
    lot.is_analyzed = False
    await db.commit()

    ctx = PipelineContext(
        tender_data={
            "title":       tender.title or "",
            "description": tender.description or "",
        },
        lot_data={
            "title":               lot.title or "",
            "description":         lot.description or "",
            "technical_spec_text": lot.technical_spec_text or "",
            "raw_spec_text":       getattr(lot, "raw_spec_text", None) or "",
        },
        tender_id=str(tender.id),
        lot_id=str(lot.id),
        platform=lot.platform or "",
    )

    await ai_analysis_step(ctx)

    return {
        "status":   "ok",
        "lot_id":   str(lot_id),
        "category": ctx.category,
        "skipped":  ctx.skip_remaining,
    }


# ── Spec text refresh helper ──────────────────────────────────────────────────

def _is_guarantee_doc(doc: dict) -> bool:
    val = " ".join([
        doc.get("name") or "",
        doc.get("url") or "",
        doc.get("row_label") or "",
    ]).lower()
    return bool(re.search(r"обеспечени|гарант|банков|template|guarantee", val))


_GUARANTEE_BODY_MARKERS = (
    "[документ: обеспечение",
    "банковская гарантия",
    "бенефициар",
    "гарантодател",
    "сумма гарантии",
    "срок действия гарантии",
    "обеспечение заявки",
)

def _looks_like_guarantee_text(text: str) -> bool:
    """
    Detect bank-guarantee templates by content (not filename). Files like
    `price_offers_guarantee_2025.docx` are caught by `_is_guarantee_doc`,
    but lots whose only document is a guarantee template are sometimes
    listed under a neutral filename (e.g. "Шаблон.docx") — content-based
    detection is the safety net so we never persist this as spec text.
    """
    if not text:
        return False
    head = text[:2000].lower()
    hits = sum(1 for m in _GUARANTEE_BODY_MARKERS if m in head)
    if hits >= 2:
        return True
    # Also catch the form-template signature: dozens of "____" placeholders +
    # the word "гарантия" / "обеспечени" anywhere in the head.
    if head.count("____") >= 5 and re.search(r"гаранти|обеспечени", head):
        return True
    return False




async def _fetch_goszakup_docs(
    external_id: str,
    client: httpx.AsyncClient,
    lot_external_id: Optional[str] = None,
) -> list[dict]:
    """
    Fetch document list from goszakup.
    Strategy:
    1. Parse the announce HTML page for direct file links.
    2. Find all actionModalShowFiles(announce_id, group) JS buttons and call
       the AJAX endpoint for each (group 125 = Техническая спецификация).
    3. From AJAX results, prefer files matching lot_external_id in the filename.
    """
    from bs4 import BeautifulSoup

    SPEC_EXTS = {".pdf", ".doc", ".docx", ".xls", ".xlsx"}
    BASE = "https://goszakup.gov.kz"
    docs: list[dict] = []
    seen: set[str] = set()

    def _parse_links(html: str, row_label: str = "") -> None:
        soup = BeautifulSoup(html, "html.parser")
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            if not href or href.startswith("#"):
                continue
            full_url = (BASE + href) if href.startswith("/") else href if href.startswith("http") else None
            if not full_url:
                continue
            path_lower = href.lower().split("?")[0]
            is_download = "download_file" in path_lower or "/files/" in path_lower
            title_attr = (a.get("title") or "").strip()
            anchor_text = a.get_text(separator=" ", strip=True)
            name = ""
            ext = ""
            for candidate in (title_attr, anchor_text, path_lower.split("/")[-1]):
                for e in SPEC_EXTS:
                    if candidate.lower().endswith(e):
                        name = candidate
                        ext = e
                        break
                if ext:
                    break
            if not ext and is_download:
                name = title_attr or anchor_text or path_lower.split("/")[-1]
                ext = ".pdf"
            if not ext:
                continue
            if full_url in seen:
                continue
            seen.add(full_url)
            label = row_label
            if not label:
                td = a.find_parent("td")
                if td:
                    tr = td.find_parent("tr")
                    if tr:
                        first_td = tr.find("td")
                        if first_td and first_td != td:
                            label = first_td.get_text(strip=True).lower()
            docs.append({"url": full_url, "name": name or full_url.split("/")[-1], "extension": ext, "is_spec": True, "row_label": label})

    # Step 1: parse the announce HTML page for direct links
    page_url = f"https://goszakup.gov.kz/ru/announce/index/{external_id}?tab=documents"
    try:
        print(f"[refresh_spec_text] fetching {page_url}", flush=True)
        resp = await client.get(page_url)
        if resp.status_code == 200:
            html = resp.text
            _parse_links(html)

            # Step 2: find actionModalShowFiles buttons and call AJAX for each
            # Techspec is group 125; contract is 101
            TECHSPEC_GROUPS = {125, 101}
            buttons = re.findall(r"actionModalShowFiles\((\d+),(\d+)\)", html)
            for ann_id_str, group_str in buttons:
                group = int(group_str)
                ajax_url = f"https://goszakup.gov.kz/ru/announce/actionAjaxModalShowFiles/{ann_id_str}/{group}"
                try:
                    ajax_resp = await client.get(
                        ajax_url,
                        headers={"X-Requested-With": "XMLHttpRequest", "Referer": page_url},
                    )
                    if ajax_resp.status_code == 200:
                        ajax_html = ajax_resp.text
                        # Parse the file table — each row: lot_number | file_link | author | org | date | sig
                        ajax_soup = BeautifulSoup(ajax_html, "html.parser")
                        for tr in ajax_soup.find_all("tr"):
                            tds = tr.find_all("td")
                            if len(tds) < 2:
                                continue
                            a_tag = tds[1].find("a", href=True)
                            if not a_tag:
                                continue
                            href = a_tag["href"].strip()
                            file_name = a_tag.get_text(strip=True)
                            full_url = (BASE + href) if href.startswith("/") else href
                            if full_url in seen:
                                continue
                            # Prefer files matching our lot's external_id
                            if lot_external_id and lot_external_id not in file_name and lot_external_id not in href:
                                # Still add, but at lower priority (append later)
                                pass
                            seen.add(full_url)
                            docs.append({
                                "url": full_url,
                                "name": file_name or full_url.split("/")[-1],
                                "extension": ".pdf",
                                "is_spec": True,
                                "row_label": "техническая спецификация" if group == 125 else "проект договора",
                                "lot_match": bool(lot_external_id and (lot_external_id in file_name or lot_external_id in href)),
                            })
                            print(f"[refresh_spec_text] AJAX group={group} file={file_name!r} url={full_url}", flush=True)
                except Exception as exc:
                    print(f"[refresh_spec_text] AJAX error group={group}: {exc}", flush=True)
    except Exception as exc:
        print(f"[refresh_spec_text] announce fetch error: {exc}", flush=True)

    # Sort: lot-matched files first, then by techspec label
    docs.sort(key=lambda d: (0 if d.get("lot_match") else 1, 0 if "техническая" in (d.get("row_label") or "") else 1))

    print(f"[refresh_spec_text] found {len(docs)} docs total", flush=True)
    for d in docs:
        print(f"  - {d['name']!r}  label={d.get('row_label')!r}  match={d.get('lot_match')}  url={d['url']}", flush=True)
    return docs


async def _refresh_spec_text(lot: TenderLot, tender: Tender, force: bool = False) -> tuple[str, str]:
    """
    Download spec PDFs from lot/tender documents (or re-fetch from goszakup if empty).
    Returns (technical_spec_text, raw_spec_text).
    Pass force=True to always re-download even if spec text already exists.
    """
    from modules.parser.document_parser import extract_text_from_bytes, truncate_for_ai, strip_kazakh_lines

    MAX_SPEC_CHARS = 10_000
    MAX_RAW_CHARS  = 50_000
    MAX_DOCS       = 3

    if not force:
        existing = (lot.technical_spec_text or "").strip()
        if len(existing) > 200:
            return existing, (getattr(lot, "raw_spec_text", None) or ""), getattr(lot, "techspec_pdf_url", None)

    raw_parts: list[str] = []
    description = (lot.description or "").strip()
    if description:
        raw_parts.append(f"[ОПИСАНИЕ ЛОТА]\n{description}")

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/pdf,*/*",
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
    }

    async with httpx.AsyncClient(timeout=40, follow_redirects=True, headers=headers) as client:
        spec_docs: list[dict] = []

        # 1) Prefer DB documents (cheap, no extra HTTP)
        all_docs: list[dict] = []
        for src in (lot.documents, tender.documents if tender else None):
            if src and isinstance(src, list):
                all_docs.extend(src)
        spec_docs = [d for d in all_docs if d.get("url") and not _is_guarantee_doc(d)]

        # 2) When force=True OR DB docs are empty: re-fetch live from goszakup
        # This is what makes auto-extract reliable for new lots whose `documents`
        # array is empty after a fresh scan.
        if (force or not spec_docs) and tender and tender.external_id and (lot.platform or "goszakup") == "goszakup":
            try:
                fetched = await _fetch_goszakup_docs(
                    tender.external_id, client, lot_external_id=lot.lot_external_id
                )
                clean = [d for d in fetched if not _is_guarantee_doc(d)]
                matched = [d for d in clean if d.get("lot_match")]
                refetched = matched[:1] if matched else clean[:1]
                if refetched:
                    spec_docs = refetched
            except Exception as exc:
                print(f"[refresh_spec_text] goszakup re-fetch error: {exc}", flush=True)

        # Cap to single best doc — extracting many is slow and rarely adds info
        spec_docs = spec_docs[:1]

        print(
            f"[refresh_spec_text] lot={lot.id} | spec_docs={len(spec_docs)} to download",
            flush=True,
        )

        pdf_url: Optional[str] = None
        for doc in spec_docs:
            url  = doc.get("url", "")
            name = doc.get("name", "") or url.split("/")[-1]
            if url.startswith("/"):
                url = "https://goszakup.gov.kz" + url
            try:
                print(f"[refresh_spec_text] GET {url}", flush=True)
                resp = await client.get(url)
                if resp.status_code != 200:
                    print(f"[refresh_spec_text] HTTP {resp.status_code} — {url}", flush=True)
                    continue
                content = resp.content
                if not content:
                    continue
                # Save PDF URL before text extraction
                if name.lower().endswith(".pdf") or b"%PDF" in content[:5]:
                    pdf_url = url
                text = extract_text_from_bytes(content, name)
                if text and len(text.strip()) > 50:
                    # Drop bank-guarantee templates that slipped past the filename
                    # filter — recognise them by content (banking placeholders +
                    # boilerplate phrases dominate). We don't want these to land
                    # in technical_spec_text and poison AI prompts / search.
                    if _looks_like_guarantee_text(text):
                        print(
                            f"[refresh_spec_text] ✗ guarantee template detected in {name!r}; skipping",
                            flush=True,
                        )
                        pdf_url = None  # don't pin a guarantee form as the spec PDF either
                        continue
                    print(f"[refresh_spec_text] ✓ {len(text)} chars from {name!r}", flush=True)
                    raw_parts.append(text)
                else:
                    print(f"[refresh_spec_text] ✗ empty/short from {name!r}", flush=True)
            except Exception as exc:
                print(f"[refresh_spec_text] error {name!r}: {exc}", flush=True)

    raw_full  = "\n\n".join(raw_parts)
    raw_full  = strip_kazakh_lines(raw_full)
    tech_text = truncate_for_ai(raw_full, MAX_SPEC_CHARS)
    return tech_text, raw_full[:MAX_RAW_CHARS], pdf_url


# ── Techspec PDF proxy ───────────────────────────────────────────────────────

@router.get("/{lot_id}/techspec-pdf")
async def proxy_techspec_pdf(
    lot_id: uuid.UUID,
    download: bool = False,
    db: AsyncSession = Depends(get_db),
):
    """Proxy the techspec PDF from goszakup so the browser can embed it."""
    lot = await db.get(TenderLot, lot_id)
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")
    pdf_url = getattr(lot, "techspec_pdf_url", None)

    # If URL not cached in DB — scan lot/tender documents for a PDF
    if not pdf_url:
        tender = await db.get(Tender, lot.tender_id) if lot.tender_id else None
        all_docs: list[dict] = []
        for src in (lot.documents, tender.documents if tender else None):
            if src and isinstance(src, list):
                all_docs.extend(src)
        for doc in all_docs:
            doc_url = doc.get("url", "")
            doc_name = (doc.get("name", "") or doc_url).lower()
            if doc_url and (doc_name.endswith(".pdf") or "pdf" in doc_url.lower()):
                if not _is_guarantee_doc(doc):
                    pdf_url = doc_url
                    # Cache it so next request is instant
                    lot.techspec_pdf_url = pdf_url
                    await db.commit()
                    break

    if not pdf_url:
        raise HTTPException(status_code=404, detail="Techspec PDF not available. Run full analysis first.")

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "application/pdf,*/*",
    }
    try:
        async with httpx.AsyncClient(timeout=30, follow_redirects=True, headers=headers) as client:
            resp = await client.get(pdf_url)
            if resp.status_code != 200:
                raise HTTPException(status_code=502, detail=f"goszakup returned {resp.status_code}")
            content = resp.content
            from urllib.parse import quote as _quote
            lot_title = (lot.title or "techspec").replace(" ", "_")[:40]
            filename_ascii = f"techspec_{lot.lot_external_id or 'lot'}.pdf"
            filename_utf8 = _quote(f"techspec_{lot_title}.pdf")
            return StreamingResponse(
                io.BytesIO(content),
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f"{'attachment' if download else 'inline'}; filename=\"{filename_ascii}\"; filename*=UTF-8''{filename_utf8}",
                    "Content-Length": str(len(content)),
                    "Cache-Control": "public, max-age=3600",
                },
            )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Failed to fetch PDF: {exc}")


# ── Full reanalysis (AI + profitability) ──────────────────────────────────────

@router.post("/{lot_id}/reanalyze-full")
async def reanalyze_lot_full(
    lot_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Full re-analysis: AI spec analysis + profitability recalculation.
    Deletes all existing analysis/profitability/supplier/logistics records for this lot.
    """
    from modules.scanner.pipeline import PipelineContext
    from modules.ai_analyzer.pipeline_step import ai_analysis_step
    from modules.profitability.engine import ProfitabilityEngine

    row = await db.execute(
        select(TenderLot, Tender)
        .join(Tender, TenderLot.tender_id == Tender.id)
        .where(TenderLot.id == lot_id)
    )
    result = row.one_or_none()
    if not result:
        raise HTTPException(status_code=404, detail="Lot not found")
    lot, tender = result

    # Clear old records
    await db.execute(sa_delete(TenderLotAnalysis).where(TenderLotAnalysis.lot_id == lot_id))
    await db.execute(sa_delete(ProfitabilityAnalysis).where(ProfitabilityAnalysis.lot_id == lot_id))

    # Clear supplier matches (and their parent suppliers that are estimates only)
    sup_rows = (await db.execute(
        select(SupplierMatch.supplier_id).where(SupplierMatch.lot_id == lot_id)
    )).scalars().all()
    await db.execute(sa_delete(SupplierMatch).where(SupplierMatch.lot_id == lot_id))
    if sup_rows:
        await db.execute(sa_delete(Supplier).where(Supplier.id.in_(sup_rows)))

    await db.execute(sa_delete(LogisticsEstimate).where(LogisticsEstimate.lot_id == lot_id))

    lot.is_analyzed = False
    lot.is_profitable = None
    lot.profit_margin_percent = None
    lot.confidence_level = None
    await db.commit()

    # Step 0: always re-download spec text from PDF (force=True clears wrong/stale content)
    tech_text, raw_text, pdf_url = await _refresh_spec_text(lot, tender, force=True)
    updated = False
    if tech_text and tech_text != (lot.technical_spec_text or ""):
        lot.technical_spec_text = tech_text
        lot.raw_spec_text = raw_text
        updated = True
    if pdf_url and pdf_url != (lot.techspec_pdf_url or ""):
        lot.techspec_pdf_url = pdf_url
        updated = True
    if updated:
        await db.commit()
        print(
            f"[reanalyze-full] spec text refreshed: {len(tech_text)} chars "
            f"(was {len(lot.technical_spec_text or '')} chars)",
            flush=True,
        )

    # Step 1: AI analysis
    ctx = PipelineContext(
        tender_data={
            "title":       tender.title or "",
            "description": tender.description or "",
        },
        lot_data={
            "title":               lot.title or "",
            "description":         lot.description or "",
            "technical_spec_text": lot.technical_spec_text or "",
            "raw_spec_text":       getattr(lot, "raw_spec_text", None) or "",
        },
        tender_id=str(tender.id),
        lot_id=str(lot.id),
        platform=lot.platform or "",
    )
    await ai_analysis_step(ctx)

    if ctx.skip_remaining or not ctx.ai_analysis:
        return {
            "status":   "skipped",
            "lot_id":   str(lot_id),
            "category": ctx.category,
            "reason":   "Category other or no AI analysis",
        }

    # Step 2: Profitability
    engine = ProfitabilityEngine()
    spec_text = lot.technical_spec_text or lot.description or ""
    prof_result = await engine.analyze(
        budget=float(lot.budget) if lot.budget else 0,
        analysis=ctx.ai_analysis,
        spec_text=spec_text,
        lot_title=lot.title or "",
        lot_id=lot_id,
        tender_id=lot.tender_id,
    )

    return {
        "status":         "ok",
        "lot_id":         str(lot_id),
        "category":       ctx.category,
        "profit_margin":  prof_result.get("profit_margin_percent") if prof_result else None,
        "is_profitable":  prof_result.get("is_profitable") if prof_result else None,
        "supplier_price_kzt": prof_result.get("product_cost") if prof_result else None,
    }


# ── Bid generation ────────────────────────────────────────────────────────────

@router.get("/{lot_id}/bid")
async def generate_lot_bid(
    lot_id: uuid.UUID,
    company_name: str = Query("Ваша компания"),
    company_bin: str = Query(""),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Generate DOCX bid proposal for a lot."""
    row = await db.execute(
        select(TenderLot, Tender)
        .join(Tender, TenderLot.tender_id == Tender.id)
        .where(TenderLot.id == lot_id)
    )
    result = row.one_or_none()
    if not result:
        raise HTTPException(status_code=404, detail="Lot not found")

    lot, tender = result

    analysis_row = await db.execute(
        select(TenderLotAnalysis).where(TenderLotAnalysis.lot_id == lot_id)
    )
    analysis = analysis_row.scalar_one_or_none()

    prof_row = await db.execute(
        select(ProfitabilityAnalysis)
        .where(ProfitabilityAnalysis.lot_id == lot_id)
        .order_by(desc(ProfitabilityAnalysis.created_at))
        .limit(1)
    )
    prof = prof_row.scalar_one_or_none()

    tender_dict = {
        "title": lot.title,
        "budget": float(lot.budget) if lot.budget else 0,
        "deadline_at": lot.deadline_at.isoformat() if lot.deadline_at else None,
        "customer_name": tender.customer_name,
    }
    analysis_dict = {}
    if analysis:
        analysis_dict = {
            "product_name": analysis.product_name,
            "technical_params": analysis.technical_params or {},
            "key_requirements": analysis.key_requirements or [],
            "summary_ru": analysis.ai_summary_ru or "",
            "spec_clarity": analysis.spec_clarity or "vague",
        }
    prof_dict = {}
    if prof:
        prof_dict = {
            "total_cost": float(prof.total_cost) if prof.total_cost else 0,
            "expected_profit": float(prof.expected_profit) if prof.expected_profit else 0,
            "profit_margin_percent": float(prof.profit_margin_percent) if prof.profit_margin_percent else 0,
            "recommended_bid": float(prof.recommended_bid) if prof.recommended_bid else 0,
            "lead_time_days": 30,
            "origin_country": prof.origin_country or "CN",
        }

    from modules.bid_generator.generator import BidProposalGenerator
    generator = BidProposalGenerator()
    docx_bytes = await generator.generate(
        tender_data=tender_dict,
        analysis=analysis_dict,
        profitability=prof_dict,
        company_name=company_name,
        company_bin=company_bin,
    )

    filename = f"bid_lot_{lot.lot_external_id}.docx"
    return StreamingResponse(
        io.BytesIO(docx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── User actions ──────────────────────────────────────────────────────────────

@router.post("/{lot_id}/action")
async def record_lot_action(
    lot_id: uuid.UUID,
    action: str = Query(..., description="viewed | ignored | bid_submitted | won | lost"),
    actual_bid_amount: Optional[float] = Query(None),
    notes: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Record user action on a lot (viewed, bid, won, lost, ignored)."""
    valid_actions = {"viewed", "ignored", "bid_submitted", "won", "lost"}
    if action not in valid_actions:
        raise HTTPException(status_code=400, detail=f"Action must be one of: {valid_actions}")

    row = await db.execute(select(TenderLot).where(TenderLot.id == lot_id))
    lot = row.scalar_one_or_none()
    if not lot:
        raise HTTPException(status_code=404, detail="Lot not found")

    user_action = UserAction(
        lot_id=lot_id,
        tender_id=lot.tender_id,
        user_id=current_user.id,
        action=action,
        actual_bid_amount=actual_bid_amount,
        notes=notes,
    )
    db.add(user_action)
    await db.commit()
    await db.refresh(user_action)

    return {"id": str(user_action.id), "action": action, "status": "recorded"}
